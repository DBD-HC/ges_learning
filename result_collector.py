import os
from openpyxl.styles import PatternFill

import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side


def merge(ws, first, last):
    ws.merge_cells(f"{first.coordinate}:{last.coordinate}")

    # 设置合并后单元格的对齐方式为居中
    merged_cell = ws.cell(row=first.row, column=first.column)
    merged_cell.border = Border(bottom=Side(style='medium'))
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

def add_acc_auc_ap(ws, row, acc_auc_aps, prefix=''):
    ws.cell(row=row + 1, column=4, value=prefix + "Acc.")
    ws.cell(row=row + 2, column=4, value=prefix + "Auc")
    cell = ws.cell(row=row + 3, column=4, value=prefix + "Ap")
    cell.border = Border(bottom=Side(style='medium'))
    for d, acc_auc_ap in enumerate(acc_auc_aps, start=5):
        ws.cell(row=row + 1, column=d, value=acc_auc_ap[0])
        ws.cell(row=row + 2, column=d, value=acc_auc_ap[1])
        cell = ws.cell(row=row + 3, column=d, value=acc_auc_ap[2])
        cell.border = Border(bottom=Side(style='medium'))


def cross_domain_results(model_name, domain, train_indexes, val_indexes, test_indexes, res, file_name='result.xlsx'):
    if not os.path.exists(file_name):
        # 创建一个工作簿对象
        wb = Workbook()
    else:
        wb = load_workbook(file_name)

    # 激活默认的工作表
    ws = wb.active
    # 确定最后一行的索引
    last_row = ws.max_row

    # 添加新行的数据
    c_title1_f = ws.cell(row=last_row + 2, column=1, value=model_name)
    c_title2_f = ws.cell(row=last_row + 2, column=2, value=','.join(map(str, train_indexes)))
    if val_indexes is None:
        val_value = 'None'
    else:
        val_value = ','.join(map(str, val_indexes))
    c_title3_f = ws.cell(row=last_row + 2, column=3, value=val_value)
    cell = ws.cell(row=last_row + 1, column=4, value=domain)
    cell.border = Border(bottom=Side(style='medium'))
    for col, value in enumerate(test_indexes, start=5):
        cell = ws.cell(row=last_row + 1, column=col, value=value)
        cell.border = Border(bottom=Side(style='medium'))

    for col in range(1, len(test_indexes) + 5):
        cell = ws.cell(row=last_row+1, column=col)
        cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type="solid")

    line_count = 1
    for t, value in enumerate(res, start=1):
        start_index = (t - 1) * 3 + 1
        add_acc_auc_ap(ws, last_row+start_index, value)
        line_count += 3

    res = np.mean(res, axis=0)
    add_acc_auc_ap(ws, last_row + line_count, res, prefix='Avg.')

    line_count = line_count + 3
    c_title1_l = ws.cell(row=last_row + line_count, column=1)
    c_title2_l = ws.cell(row=last_row + line_count, column=2)
    c_title3_l = ws.cell(row=last_row + line_count, column=3)

    merge(ws, c_title1_f, c_title1_l)
    merge(ws, c_title2_f, c_title2_l)
    merge(ws, c_title3_f, c_title3_l)

    wb.save(file_name)


if __name__ == '__main__':
    res = np.random.random((5, 5, 3))

    cross_domain_results(model_name='xx',
                         domain=1,
                         train_indexes=[0, 1, 2],
                         val_indexes=[0, 1, 2],
                         test_indexes=[0, 1, 2, 3, 4],
                         res=res,
                         file_name='result.xlsx')

